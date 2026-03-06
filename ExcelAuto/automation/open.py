#Open an existing woorkbook
from openpyxl import load_workbook

#load an existing workbook
wb = load_workbook("employees.xlsx")

#view sheet names
print(wb.sheetnames)

#select an active sheet
sheet = wb.active
print("Active sheet:", sheet.title)

#Reading cell values, rows and columns
#Reading a single cell
value = sheet["A1"].value
print("A1", value)

#Reading using cell() method
value = sheet.cell(row=2, column=1).value
print("Row 2, Col 1:", value)

#reading multiple cells
#read all values in column A
for cell in sheet["A"]:
    print(cell.value)

#read all values in row 1
for cell in sheet[1]:
    print(cell.value)

#Reading a Range
for row in sheet["A1:B3"]:
    for cell in row:
        print(cell.value, end = " ")
    print()

# #writing data into cells
sheet["D1"] =  "Grade"
sheet.cell(2, 4).value = "A"
sheet.cell(3, 4).value = "B+"
sheet.cell(4, 4).value = "A+"

#range after adding new cells
for row in sheet["A1:D4"]:
    for cell in row:
        print(cell.value, end = " ")
    print()

# saving and renaming workbooks: you must save workbooks after making changes
wb.save("employees.xlsx")
print("Workbook saved successfully")
