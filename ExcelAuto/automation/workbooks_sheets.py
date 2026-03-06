#Creating a new workbook and worksheet
#creating a new workbook
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
sheet = wb.active
sheet.title = "workers"

sheet["A1"] = "Name"
sheet["B1"] = "Salary"

wb.save("workers.xlsx")
print("New workbook created!")

#Adding a new worksheet
wb = load_workbook("workers.xlsx")
sheet = wb.create_sheet("Bonus")
print(sheet.title)

# save the update
wb.save("workers.xlsx")
print(wb.worksheets) #confirm the number of sheets created

#reference to 1 worksheet
sheet1=wb["workers"]
sheet2=wb["Bonus"]
print(sheet1.title)
print(sheet2.title)

#renaming a sheet
sheet = wb["Bonus"]
sheet.title = "salary"
wb.save("workers.xlsx")
print(wb.worksheets)

#deleting a sheet
del wb["salary"]
wb.save("workers.xlsx")
print("sheet deleted!")
print(wb.worksheets)
